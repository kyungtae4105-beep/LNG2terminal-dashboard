"""
월간안전공정회의 엑셀 → schedule_data.json/.js 생성.

추출 시트:
- 3.Master Schedule_Rev.0 / Rev.3
- 3-1.Master Schedule_Rev.6B
- 3-1.Updated Schedule (Master Rev.4 / 공정만회 / Updated 3개 비교)
- 3-2.Process 상세 공정표_Rev.2B (Rev.1 vs Rev.2B 비교 + 부모-자식 트리)
- 4.월간시공계획(3개월) (계획/변경/실적 3개 비교)

Process 시트의 들여쓰기 규칙:
  - B 컬럼 텍스트 → 그룹 헤더, 앞 공백 갯수 = level
  - C 컬럼 텍스트 → 작업/하위그룹
  - parent_idx 를 같이 부여하여 트리 구조 형성
"""
import os
import sys
import json
import datetime as dt
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

SOURCE_FILE = "광)LNG #7,8탱크_월간안전공정회의_26.5월_Final.xlsx"
OUTPUT_FILE = "schedule_data.json"


def to_iso(v):
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    return None


def cell(ws, r, col):
    return ws.cell(row=r, column=openpyxl.utils.column_index_from_string(col)).value


def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


# ---------- Master Schedule (Rev.0 / Rev.3 / Rev.6B) ----------
def extract_master(ws, *, name_cols, start_col, end_col, dur_col, first_row, last_row=None):
    """name_cols : 들여쓰기 단계별 컬럼 리스트(좌→우, 상→하)"""
    last_row = last_row or ws.max_row
    tasks = []
    for r in range(first_row, last_row + 1):
        cells = {c: cell(ws, r, c) for c in (list(name_cols) + [start_col, end_col, dur_col])}
        # 어느 name_col이 채워졌는지로 level 결정
        nm_idx = None
        nm_val = None
        for i, c in enumerate(name_cols):
            v = cells.get(c)
            if isinstance(v, str) and v.strip():
                nm_idx, nm_val = i, v.strip()
        s = to_iso(cells.get(start_col))
        e = to_iso(cells.get(end_col))
        d = cells.get(dur_col)
        if not isinstance(d, (int, float)):
            d = None
        if nm_val is None and s is None and e is None:
            continue
        tasks.append({
            "level": nm_idx or 0,
            "name": nm_val or "",
            "start": s,
            "end": e,
            "duration_days": int(d) if d is not None else None,
        })
    # parent_idx 부여 — 자기보다 상위 레벨의 가장 가까운 행
    assign_parents(tasks)
    return tasks


def assign_parents(tasks):
    """tasks 리스트에 parent_idx 부여 (level 기반 스택)."""
    stack = []  # (level, idx)
    for i, t in enumerate(tasks):
        lv = t.get("level", 0)
        while stack and stack[-1][0] >= lv:
            stack.pop()
        t["parent_idx"] = stack[-1][1] if stack else None
        stack.append((lv, i))


# ---------- Updated Schedule (3개 비교) ----------
def extract_updated(ws):
    name_cols = ["A", "B", "C", "D"]
    out = []
    for r in range(7, ws.max_row + 1):
        cells = {c: cell(ws, r, c) for c in name_cols + ["E", "F", "G", "H", "I", "J", "K", "L", "M"]}
        nm_idx, nm_val = None, None
        for i, c in enumerate(name_cols):
            v = cells.get(c)
            if isinstance(v, str) and v.strip():
                nm_idx, nm_val = i, v.strip()
        rec = {
            "level": nm_idx or 0,
            "name": nm_val or "",
            "rev4":     mkrev(cells["E"], cells["F"], cells["G"]),
            "recovery": mkrev(cells["H"], cells["I"], cells["J"]),
            "updated":  mkrev(cells["K"], cells["L"], cells["M"]),
        }
        if rec["name"] == "" and rec["rev4"]["start"] is None and rec["recovery"]["start"] is None and rec["updated"]["start"] is None:
            continue
        out.append(rec)
    assign_parents(out)
    return out


def mkrev(s, e, d):
    return {"start": to_iso(s), "end": to_iso(e), "duration_days": int(d) if isinstance(d, (int, float)) else None}


# ---------- Process 상세 공정표 ----------
def extract_process(ws):
    """
    B 컬럼 = 그룹 헤더 (앞 공백 = level)
    C 컬럼 = 작업/하위 그룹 (마지막 B 헤더 level + 1)
            텍스트가 '-'로 시작하면 sub-header (작업 level + 0), 아니면 task (+1)
    """
    out = []
    for r in range(9, ws.max_row + 1):
        cA = cell(ws, r, "A")
        cB = cell(ws, r, "B")
        cC = cell(ws, r, "C")
        f, g, h = cell(ws, r, "F"), cell(ws, r, "G"), cell(ws, r, "H")
        i, j, k = cell(ws, r, "I"), cell(ws, r, "J"), cell(ws, r, "K")
        rev1 = mkrev(f, g, h)
        rev2b = mkrev(i, j, k)

        name = None
        kind = None  # 'header' | 'subheader' | 'task'
        level = None
        if isinstance(cB, str) and cB.strip():
            level = len(cB) - len(cB.lstrip())
            name = cB.strip()
            kind = "header"
        elif isinstance(cC, str) and cC.strip():
            name = cC.strip()
            kind = "subheader" if name.startswith("-") else "task"
            # 가장 최근에 본 B 헤더의 level 사용
            # 마지막 outut에서 kind=='header'인 것의 level + 1
            last_hdr_lv = None
            for prev in reversed(out):
                if prev.get("kind") == "header":
                    last_hdr_lv = prev["level"]
                    break
            base = (last_hdr_lv if last_hdr_lv is not None else 0) + 1
            level = base if kind == "subheader" else base + 1

        if name is None and rev1["start"] is None and rev2b["start"] is None:
            continue

        wbs = cA if isinstance(cA, (int, float, str)) and str(cA).strip() else None

        out.append({
            "level": level if level is not None else 0,
            "kind": kind or "task",
            "wbs": wbs,
            "name": name or "",
            "rev1": rev1,
            "rev2b": rev2b,
        })

    assign_parents(out)
    return out


# ---------- 월간시공계획 (3개월) ----------
def extract_monthly(ws):
    """
    B/C/D 컬럼 = 공종 계층
    E 컬럼 = 계획/변경/실적
    F/G  = 시작일/종료일

    같은 (B,C,D) 묶음 안에서 E=계획/변경/실적이 3행씩 나오므로,
    하나의 task로 묶고 plan/change/actual 3개 막대로 만든다.
    """
    tasks = []
    cur_path = ["", "", ""]  # B, C, D 최신 값
    # 작업을 (B,C,D) key 로 인덱싱
    by_key = {}
    for r in range(8, ws.max_row + 1):
        bv = cell(ws, r, "B")
        cv = cell(ws, r, "C")
        dv = cell(ws, r, "D")
        ev = cell(ws, r, "E")
        fv = cell(ws, r, "F")
        gv = cell(ws, r, "G")

        if isinstance(bv, str) and bv.strip():
            cur_path[0] = bv.strip()
            cur_path[1] = ""
            cur_path[2] = ""
        if isinstance(cv, str) and cv.strip():
            cur_path[1] = cv.strip().replace("\n", " ")
            cur_path[2] = ""
        if isinstance(dv, str) and dv.strip():
            cur_path[2] = dv.strip()

        if not isinstance(ev, str) or ev.strip() not in ("계획", "변경", "실적"):
            continue
        kind = ev.strip()
        key = tuple(cur_path)
        if key not in by_key:
            t = {
                "category": cur_path[0],
                "subgroup": cur_path[1],
                "activity": cur_path[2],
                "plan":   {"start": None, "end": None},
                "change": {"start": None, "end": None},
                "actual": {"start": None, "end": None},
            }
            by_key[key] = t
            tasks.append(t)
        t = by_key[key]
        slot = {"계획": "plan", "변경": "change", "실적": "actual"}[kind]
        t[slot] = {"start": to_iso(fv), "end": to_iso(gv)}
    return tasks


# ---------- 10.기자재발주현황 ----------
def extract_orders(ws):
    """
    각 LOT 는 두 행(Plan 행 + Actual 행)으로 구성.
    B(No.) 가 정수인 행이 LOT 시작 = Plan 행.
    그 다음 행 (P='A') 이 Actual 행.
    """
    # 컬럼 매핑
    COLS = {
        'no': 'B', 'pec_kind': 'C', 'discipline': 'D',
        'desc': 'E', 'status_text': 'F', 'vendor_raw': 'G',
        'pm_sales': 'H', 'pm_design': 'I',
        'delivery_period': 'J', 'vendor_sel': 'K',
        'item_no': 'L', 'item_name': 'M',
        'qty': 'N', 'unit': 'O',
        'pa': 'P',
        'rfq_issue': 'Q', 'rfq_approval': 'R',
        'tbe_issue': 'S', 'tbe_approval': 'T',
        'po': 'U', 'key_vp': 'V',
        'fob': 'W', 'on_site': 'X',
        'remarks': 'Y',
    }
    # 병합셀로 빈 셀 채우기 위한 마지막 값 보관
    last_pec = ''
    last_discipline = ''
    rows = []
    r = 4
    max_r = ws.max_row
    while r <= max_r:
        nv = cell(ws, r, 'B')
        if not isinstance(nv, int):
            r += 1
            continue
        # P 행 (계획)
        plan = {k: cell(ws, r, c) for k, c in COLS.items()}
        # 다음 A 행이 있는지 확인 (P='A')
        if r + 1 <= max_r and cell(ws, r + 1, 'P') == 'A':
            actual = {k: cell(ws, r + 1, c) for k, c in COLS.items()}
            next_r = r + 2
        else:
            actual = {}
            next_r = r + 1
        # 병합 셀 보완
        pec = plan.get('pec_kind') or last_pec
        if pec: last_pec = pec
        disc = plan.get('discipline') or last_discipline
        if disc: last_discipline = disc
        # 정규화
        def _d(obj, k):
            v = obj.get(k) if isinstance(obj, dict) else None
            return to_iso(v)
        def _str(v):
            return str(v).strip() if v is not None else ''
        lot = {
            'lotNo': int(nv),
            'pec': _str(pec).replace('\n', ' '),
            'discipline': _str(disc),
            'itemName': _str(plan.get('desc')).replace('\n', ' '),
            'itemNameDetail': _str(plan.get('item_name')).replace('\n', ' '),
            'vendor': _str(plan.get('vendor_sel') or plan.get('vendor_raw')).replace('\n', ' '),
            'pmSales': _str(plan.get('pm_sales')),
            'pmDesign': _str(plan.get('pm_design')),
            'deliveryPeriod': _str(plan.get('delivery_period')),
            'qty': _str(plan.get('qty')),
            'unit': _str(plan.get('unit')),
            'rfqIssuePlan':   _d(plan, 'rfq_issue'),
            'rfqIssueActual': _d(actual, 'rfq_issue'),
            'rfqApprovalPlan':   _d(plan, 'rfq_approval'),
            'rfqApprovalActual': _d(actual, 'rfq_approval'),
            'tbeIssuePlan':   _d(plan, 'tbe_issue'),
            'tbeIssueActual': _d(actual, 'tbe_issue'),
            'tbeApprovalPlan':   _d(plan, 'tbe_approval'),
            'tbeApprovalActual': _d(actual, 'tbe_approval'),
            'poPlan':   _d(plan, 'po'),
            'poActual': _d(actual, 'po'),
            'keyVpPlan':   _d(plan, 'key_vp'),
            'keyVpActual': _d(actual, 'key_vp'),
            'fobPlan':   _d(plan, 'fob'),
            'fobActual': _d(actual, 'fob'),
            'onSitePlan':   _d(plan, 'on_site'),
            'onSiteActual': _d(actual, 'on_site'),
            'statusText': _str(plan.get('status_text')).replace('\n', ' '),
            'remarks':    _str(plan.get('remarks')).replace('\n', ' '),
        }
        # 단계(stage) 자동 판단
        if lot['onSiteActual']:
            lot['stage'] = 'OnSite'
        elif lot['fobActual'] or (lot['fobPlan'] and isinstance(plan.get('fob'), str) and plan.get('fob').upper() != 'N/A'):
            lot['stage'] = 'FOB'
        elif lot['poActual']:
            lot['stage'] = '제작'
        elif lot['tbeApprovalActual']:
            lot['stage'] = 'PO'
        elif lot['tbeIssueActual'] or lot['rfqApprovalActual']:
            lot['stage'] = 'TBE'
        elif lot['rfqIssueActual']:
            lot['stage'] = 'RFQ'
        else:
            lot['stage'] = '미발행'
        # 지연일 (onSitePlan vs onSiteActual)
        if lot['onSitePlan'] and lot['onSiteActual']:
            from datetime import date as _date
            try:
                p = _date.fromisoformat(lot['onSitePlan'])
                a = _date.fromisoformat(lot['onSiteActual'])
                lot['delayDays'] = (a - p).days
            except Exception:
                lot['delayDays'] = None
        elif lot['onSitePlan'] and not lot['onSiteActual']:
            from datetime import date as _date
            try:
                p = _date.fromisoformat(lot['onSitePlan'])
                today = _date.today()
                if today > p:
                    lot['delayDays'] = (today - p).days
                else:
                    lot['delayDays'] = 0
            except Exception:
                lot['delayDays'] = None
        else:
            lot['delayDays'] = None
        # 리드타임 (RFQ Plan → PO Actual)
        if lot['rfqIssuePlan'] and lot['poActual']:
            from datetime import date as _date
            try:
                rfq = _date.fromisoformat(lot['rfqIssuePlan'])
                po = _date.fromisoformat(lot['poActual'])
                lot['leadTimeDays'] = (po - rfq).days
            except Exception:
                lot['leadTimeDays'] = None
        else:
            lot['leadTimeDays'] = None
        rows.append(lot)
        r = next_r
    return rows


# ---------- 8.설계진도현황 ----------
def extract_design(ws):
    """
    분야별:
      B = 분야명
      C = 보할 (A)
      D = 실적 진도 (B)
      E = 실적율 (B/A)
      F/G/H = 승인용 (전기간) 계획/실적/잔여
      I/J/K = 공사용 (전기간) 계획/실적/잔여
      L/M/N = 승인용 (월말기준) 계획/실적/잔여
      O/P/Q = 공사용 (월말기준) 계획/실적/잔여
      R = 실적진도율 (계획대비)
    """
    fields = []
    base_ym = None
    # 헤더에서 기준월 추출 (D3 = '′25년 12월\n실적 진도\n(B)')
    d3 = cell(ws, 3, "D")
    if isinstance(d3, str):
        import re as _re
        m = _re.search(r"['′]?(\d{2,4})\s*년\s*(\d{1,2})", d3)
        if m:
            y = int(m.group(1))
            if y < 100:
                y += 2000
            base_ym = f"{y:04d}-{int(m.group(2)):02d}"
    for r in range(6, ws.max_row + 1):
        bv = cell(ws, r, "B")
        if not isinstance(bv, str) or not bv.strip():
            continue
        name = bv.strip()
        cv = cell(ws, r, "C")  # 보할
        dv = cell(ws, r, "D")  # 실적
        ev = cell(ws, r, "E")  # 실적율
        if not isinstance(cv, (int, float)) and not isinstance(dv, (int, float)):
            continue
        def _num(v):
            return float(v) if isinstance(v, (int, float)) else None
        rec = {
            "name": name,
            "weight":     _num(cv),
            "actual":     _num(dv),
            "ratio":      _num(ev),
            "approval_total":  {"plan": _num(cell(ws, r, "F")), "actual": _num(cell(ws, r, "G")), "remain": _num(cell(ws, r, "H"))},
            "construction_total": {"plan": _num(cell(ws, r, "I")), "actual": _num(cell(ws, r, "J")), "remain": _num(cell(ws, r, "K"))},
            "approval_month":  {"plan": _num(cell(ws, r, "L")), "actual": _num(cell(ws, r, "M")), "remain": _num(cell(ws, r, "N"))},
            "construction_month": {"plan": _num(cell(ws, r, "O")), "actual": _num(cell(ws, r, "P")), "remain": _num(cell(ws, r, "Q"))},
            "plan_ratio": _num(cell(ws, r, "R")),
        }
        fields.append(rec)
    # 합계 (마지막 행 — B가 비어 있고 C/D는 합)
    summary = None
    for r in range(ws.max_row, 5, -1):
        if isinstance(cell(ws, r, "B"), str) and cell(ws, r, "B").strip():
            continue
        cv = cell(ws, r, "C")
        dv = cell(ws, r, "D")
        if isinstance(cv, (int, float)) and isinstance(dv, (int, float)):
            summary = {"weight": float(cv), "actual": float(dv), "ratio": float(dv) / float(cv) if cv else None}
            break
    return {"as_of": base_ym, "fields": fields, "summary": summary}


# ---------- 5-1.물량처리계획 (EPC 단계별 공정률) ----------
def extract_qty_progress(ws):
    """
    B/C 열의 EPC 단계와 세부 공종.
    P열 = 5월 계획 EPC, T열 = 5월 예상 EPC, I열 = 4월 계획 EPC, L열 = 4월 실적 EPC
    F열 = 가중치
    상위 EPC = ENGINEERING / PROCUREMENT / CONSTRUCTION
    """
    def _num(v):
        return float(v) if isinstance(v, (int, float)) else None
    rows = []
    for r in range(7, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        f_v = ws.cell(row=r, column=6).value
        p_v = ws.cell(row=r, column=16).value  # 5월 계획 EPC
        t_v = ws.cell(row=r, column=20).value  # 5월 예상 EPC
        i_v = ws.cell(row=r, column=9).value   # 4월 계획 EPC
        l_v = ws.cell(row=r, column=12).value  # 4월 실적 EPC
        name = None
        kind = None
        if isinstance(b, str) and b.strip():
            name = b.strip()
            kind = 'group' if name.isupper() or name in ('종합 공정률 (EPC)',) else 'sub'
        elif isinstance(c, str) and c.strip():
            name = c.strip()
            kind = 'leaf'
        if name is None:
            continue
        rows.append({
            'name': name,
            'kind': kind,
            'weight': _num(f_v),
            'plan_apr': _num(i_v),
            'actual_apr': _num(l_v),
            'plan_may': _num(p_v),
            'expect_may': _num(t_v),
        })
    return rows


# ---------- 5-2.인원동원계획 ----------
def extract_manpower(ws):
    """
    5번째 행부터 9번째 행에 공종별 합계 + 합계 행.
    B5: 토목공사, B6: 건축공사, B7: 기계공사, B8: 전기/계장공사, B9: 합계
    컬럼: C=전월실적누계, D=4월실적, E=4월일평균, F=4월실적누계,
          G=5월계획, H=5월일평균
    """
    def _num(v):
        return float(v) if isinstance(v, (int, float)) else None
    out = {'categories': [], 'total': None}
    for r in range(5, 10):
        nm = ws.cell(row=r, column=2).value
        if not isinstance(nm, str): continue
        nm = nm.strip()
        rec = {
            'name': nm,
            'prev_cum':   _num(ws.cell(row=r, column=3).value),
            'actual_apr': _num(ws.cell(row=r, column=4).value),
            'daily_apr':  _num(ws.cell(row=r, column=5).value),
            'cum_apr':    _num(ws.cell(row=r, column=6).value),
            'plan_may':   _num(ws.cell(row=r, column=7).value),
            'daily_may':  _num(ws.cell(row=r, column=8).value),
        }
        if nm.replace(' ', '') == '합계':
            out['total'] = rec
        else:
            out['categories'].append(rec)
    return out


# ---------- 5-3.장비동원계획 ----------
def extract_equipment(ws):
    """
    공종별 sub-table 4개 (토목/건축/기계/전기계장). 각 sub-table:
    - 헤더 행 (B=N. XX공사) + 2행 후 첫 장비 행
    - 장비명 B열, 4월 실적 N열, 총계 O열, 5월 일평균 계획 Q열
    - 같은 sub-table 안에서 마지막 합계 행은 B에 공종명만 (예: '토목공사')
    같은 장비를 4공종 합산.
    """
    def _num(v):
        return float(v) if isinstance(v, (int, float)) else 0
    # 공종 헤더 찾기
    section_heads = []  # (header_row, section_name)
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if isinstance(b, str) and b.strip() and b.strip()[:2] in ('1.', '2.', '3.', '4.', '5.'):
            section_heads.append((r, b.strip()))
    sections = []  # 각 공종 별
    agg = {}  # equipment name -> {actual_apr, plan_may}
    for idx, (hr, hn) in enumerate(section_heads):
        next_hr = section_heads[idx + 1][0] if idx + 1 < len(section_heads) else ws.max_row + 1
        sec_name = hn.split('.', 1)[1].strip()
        sec_data = {'name': sec_name, 'items': []}
        # 장비 행: hr+3 ~ next_hr-2 (마지막은 합계)
        for r in range(hr + 3, next_hr - 1):
            eq = ws.cell(row=r, column=2).value
            if not isinstance(eq, str) or not eq.strip(): continue
            eq = eq.strip()
            # 합계행은 sec_name과 같음
            if eq == sec_name: continue
            actual_apr = _num(ws.cell(row=r, column=14).value)  # N
            total_cum  = _num(ws.cell(row=r, column=15).value)  # O
            plan_may   = _num(ws.cell(row=r, column=17).value)  # Q (일평균 계획)
            rec = {'name': eq, 'actual_apr': actual_apr, 'total_cum': total_cum, 'plan_may_daily': plan_may}
            sec_data['items'].append(rec)
            if eq not in agg:
                agg[eq] = {'name': eq, 'actual_apr': 0, 'total_cum': 0, 'plan_may_daily': 0}
            agg[eq]['actual_apr'] += actual_apr
            agg[eq]['total_cum'] += total_cum
            agg[eq]['plan_may_daily'] += plan_may
        sections.append(sec_data)
    # Top 정렬
    aggregated = sorted(agg.values(), key=lambda x: x['actual_apr'], reverse=True)
    return {'sections': sections, 'aggregated': aggregated}


# ---------- 모든 월간회의 엑셀에서 EPC 시계열 추출 ----------
def extract_epc_series_all():
    """폴더 내 모든 '광*월간안전공정회의*.xlsx'에서 5-1 시트의 종합/설계/구매/시공 누계 EPC를 월별 시계열로 추출."""
    import glob, re
    def safe_num(v):
        return float(v) if isinstance(v, (int, float)) else None
    def find_row_by_name(ws, name):
        for r in range(7, 60):
            v = ws.cell(row=r, column=2).value
            if isinstance(v, str) and v.strip() == name:
                return r
        return None
    series = []
    files = sorted(glob.glob("광*월간안전공정회의*.xlsx") +
                   glob.glob("터미널건설추진반*월간안전공정회의*.xlsx"))
    for fp in files:
        m = re.search(r"(\d{2})\.(\d{1,2})월", fp)
        if not m: continue
        year = 2000 + int(m.group(1))
        month = int(m.group(2))
        ym = f"{year:04d}-{month:02d}"
        try:
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        except Exception as e:
            print(f"  skip {fp[:50]}: {e}")
            continue
        if "5-1.물량처리계획" not in wb.sheetnames:
            wb.close(); continue
        ws = wb["5-1.물량처리계획"]
        epc_r  = find_row_by_name(ws, "종합 공정률 (EPC)")
        eng_r  = find_row_by_name(ws, "ENGINEERING")
        proc_r = find_row_by_name(ws, "PROCUREMENT")
        cons_r = find_row_by_name(ws, "CONSTRUCTION")
        def grab(rn):
            if not rn: return None
            return {
                "plan":   safe_num(ws.cell(row=rn, column=16).value),  # P: 보고월 계획
                "actual": safe_num(ws.cell(row=rn, column=20).value),  # T: 보고월 예상실적
            }
        rec = {
            "ym": ym, "source": fp,
            "epc": grab(epc_r), "eng": grab(eng_r),
            "proc": grab(proc_r), "cons": grab(cons_r),
        }
        # ym 중복 시 더 큰 actual로 (R 버전 우선)
        existing = next((s for s in series if s["ym"] == ym), None)
        if existing:
            ea = existing["epc"]["actual"] if existing["epc"] else 0
            na = rec["epc"]["actual"] if rec["epc"] else 0
            if (na or 0) >= (ea or 0):
                series.remove(existing); series.append(rec)
        else:
            series.append(rec)
        wb.close()
    series.sort(key=lambda r: r["ym"])
    return series


def main():
    if not os.path.exists(SOURCE_FILE):
        sys.exit(f"Source file not found: {SOURCE_FILE}")
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True, read_only=False)

    result = {
        "source_file": SOURCE_FILE,
        "extracted_at": dt.date.today().isoformat(),
        "master_schedule": {},
        "process_schedule": None,
        "monthly_schedule": None,
        "design_progress": None,
        "orders": None,
        "qty_progress": None,
        "manpower_may": None,
        "equipment_may": None,
        "epc_series": None,
    }

    if "3.Master Schedule_Rev.3" in wb.sheetnames:
        ws = wb["3.Master Schedule_Rev.3"]
        result["master_schedule"]["rev3"] = {
            "title": "Master Schedule Rev.3 (1차변경계약)",
            "period": "2022.10.25 ~ 2026.07.31",
            "tasks": extract_master(ws, name_cols=["A", "B", "C", "D"], start_col="E", end_col="F", dur_col="G", first_row=9),
        }

    if "3.Master Schedule_Rev.0" in wb.sheetnames:
        ws = wb["3.Master Schedule_Rev.0"]
        result["master_schedule"]["rev0"] = {
            "title": "Master Schedule Rev.0 (최초계약)",
            "period": None,
            "tasks": extract_master(ws, name_cols=["B", "C", "D"], start_col="E", end_col="F", dur_col="G", first_row=7),
        }

    if "3-1.Master Schedule_Rev.6B" in wb.sheetnames:
        ws = wb["3-1.Master Schedule_Rev.6B"]
        result["master_schedule"]["rev6b"] = {
            "title": "Master Schedule Rev.6B (최종변경계약, 승인전)",
            "period": "2022.10.25 ~ 2026.10.31",
            "tasks": extract_master(ws, name_cols=["B", "C", "D", "E"], start_col="F", end_col="G", dur_col="H", first_row=8),
        }

    if "3-1.Updated Schedule" in wb.sheetnames:
        ws = wb["3-1.Updated Schedule"]
        result["master_schedule"]["updated"] = {
            "title": "Updated Schedule (2024.07.02 공정만회)",
            "period": None,
            "compare": ["rev4", "recovery", "updated"],
            "tasks": extract_updated(ws),
        }

    if "3-2.Process 상세 공정표_Rev.2B" in wb.sheetnames:
        ws = wb["3-2.Process 상세 공정표_Rev.2B"]
        result["process_schedule"] = {
            "title": "Process 상세 공정표 Rev.2B",
            "compare": ["rev1", "rev2b"],
            "tasks": extract_process(ws),
        }

    if "8.설계진도현황" in wb.sheetnames:
        ws = wb["8.설계진도현황"]
        result["design_progress"] = extract_design(ws)

    if "4.월간시공계획(3개월)" in wb.sheetnames:
        ws = wb["4.월간시공계획(3개월)"]
        result["monthly_schedule"] = {
            "title": "월간시공계획 (3개월)",
            "compare": ["plan", "change", "actual"],
            "tasks": extract_monthly(ws),
        }

    # 10.기자재발주현황 — 시트명 변형 대응
    orders_sheet = None
    for sn in wb.sheetnames:
        if "기자재" in sn and "발주" in sn:
            orders_sheet = sn
            break
    if orders_sheet:
        ws = wb[orders_sheet]
        result["orders"] = {
            "title": "기자재 발주현황",
            "sheet": orders_sheet,
            "lots": extract_orders(ws),
        }

    if "5-1.물량처리계획" in wb.sheetnames:
        result["qty_progress"] = {
            "title": "5-1. 물량처리계획 (EPC 공정률)",
            "rows": extract_qty_progress(wb["5-1.물량처리계획"]),
        }
    if "5-2.인원동원계획" in wb.sheetnames:
        result["manpower_may"] = {
            "title": "5-2. 인원 동원 (5월 계획 + 4월 실적)",
            **extract_manpower(wb["5-2.인원동원계획"]),
        }
    if "5-3.장비동원계획" in wb.sheetnames:
        result["equipment_may"] = {
            "title": "5-3. 장비 동원 (4월 실적 + 5월 일평균 계획)",
            **extract_equipment(wb["5-3.장비동원계획"]),
        }

    wb.close()

    # 모든 월간회의 엑셀에서 EPC 시계열 추출
    result["epc_series"] = {
        "title": "EPC 누계 시계열 (월간회의 엑셀 전체)",
        "rows": extract_epc_series_all(),
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    js_path = OUTPUT_FILE.replace(".json", ".js")
    with open(js_path, "w", encoding="utf-8") as f:
        f.write("window.SCHEDULE_DATA = ")
        json.dump(result, f, ensure_ascii=False)
        f.write(";\n")

    print(f"✓ Saved {OUTPUT_FILE} + {js_path}")
    for k, v in result["master_schedule"].items():
        print(f"  master_schedule.{k}: {len(v['tasks'])} rows")
    if result["process_schedule"]:
        print(f"  process_schedule: {len(result['process_schedule']['tasks'])} rows")
    if result["monthly_schedule"]:
        print(f"  monthly_schedule: {len(result['monthly_schedule']['tasks'])} rows")
    if result["design_progress"]:
        print(f"  design_progress: {len(result['design_progress']['fields'])} fields, as_of={result['design_progress']['as_of']}")
    if result["orders"]:
        print(f"  orders: {len(result['orders']['lots'])} lots (sheet={result['orders']['sheet']})")
    if result["qty_progress"]:
        print(f"  qty_progress: {len(result['qty_progress']['rows'])} rows")
    if result["manpower_may"]:
        cats = result["manpower_may"].get("categories", [])
        total = result["manpower_may"].get("total")
        print(f"  manpower_may: {len(cats)} categories · 5월 계획 합계 {total['plan_may'] if total else '?'}")
    if result["equipment_may"]:
        agg = result["equipment_may"].get("aggregated", [])
        top = agg[:3]
        print(f"  equipment_may: {len(agg)} unique equipments · Top3 4월실적 {[(t['name'], t['actual_apr']) for t in top]}")
    if result["epc_series"]:
        rs = result["epc_series"]["rows"]
        print(f"  epc_series: {len(rs)} months from {rs[0]['ym'] if rs else '-'} to {rs[-1]['ym'] if rs else '-'}")


if __name__ == "__main__":
    main()
