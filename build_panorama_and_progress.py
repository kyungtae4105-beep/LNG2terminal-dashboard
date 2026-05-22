"""
build_panorama_and_progress.py

1) photos.js 에서 '전경' 캡션 중 월별 대표 1장(육상/현장부 위주)을 골라
   panorama_data.js (window.PANORAMA_DATA) 로 저장
2) 26.5월 엑셀의 '2-1.공정현황' 시트를 파싱해 progress_status_data.js
   (window.PROGRESS_STATUS_DATA) 로 저장
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
PHOTOS_JS = ROOT / "photos.js"
EXCEL = ROOT / "광)LNG #7,8탱크_월간안전공정회의_26.5월_Final.xlsx"

OUT_PANO = ROOT / "panorama_data.js"
OUT_PROG = ROOT / "progress_status_data.js"
OUT_EVENT = ROOT / "events_data.js"
OUT_QUAL = ROOT / "quality_data.js"

# ---- 1) 전경 사진 ----------------------------------------------------------

def load_photos():
    """photos.js → dict (window.PHOTOS_DATA 값 그대로)"""
    text = PHOTOS_JS.read_text(encoding="utf-8")
    # strip 'window.PHOTOS_DATA = ' prefix and trailing ';'
    text = re.sub(r"^\s*window\.PHOTOS_DATA\s*=\s*", "", text, count=1)
    text = text.rstrip().rstrip(";")
    return json.loads(text)


# 캡션 패턴 우선순위 — 가능한 한 '육상/현장 전체'를 찍은 컷
# (먼저 매칭되는 패턴이 그 월의 대표 사진)
PRIORITY_PATTERNS = [
    r"제2\s*LNG\s*터미널\s*육상부\s*전경",      # 2025.12 이후
    r"제2\s*LNG\s*터미널\s*현장\s*전경",         # 2024.07 ~ 2025.11
    r"#7,?\s*8\s*탱크\s*외부\s*전경",            # 2023.06 ~ 2024.06
    r"#6\s*탱크\s*외부\s*전경",                  # 2023.02 ~ 2023.11
    r"현장\s*전경",                              # fallback
]


def pick_panorama(items):
    """items 중 우선순위 패턴에 맞는 첫 1장 반환"""
    for pat in PRIORITY_PATTERNS:
        rgx = re.compile(pat)
        for it in items:
            cap = it.get("caption") or ""
            if rgx.search(cap):
                return it
    return None


MONTH_RGX = re.compile(r"^([0-9]{4})-([0-9]{2})$|^([0-9]{2})월$")


def normalize_period(key):
    """'2023-02' / '01월' 같은 키를 (year, month, ym_label) 로 정규화"""
    m = re.match(r"^(\d{4})-(\d{2})$", key)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        return y, mo, f"{y}-{mo:02d}"
    m = re.match(r"^(\d{2})월$", key)
    if m:
        # 연도 단서 없는 '01월~12월' 폴더 — 현재 2025/2026 로 추정
        # photos.js 의 폴더 명을 보면 25.06 이전은 2025-MM, 이후는 2026-MM 패턴
        # 'NN월' 폴더는 25년 1~5월 시점에 채집된 컷 → 2025년으로 간주
        mo = int(m.group(1))
        y = 2025
        return y, mo, f"{y}-{mo:02d}"
    return None


PROCESS_TYPES = {"공정사진", "공사사진"}
LAND_CATS = {"육상부"}


def build_panorama():
    photos = load_photos()
    selected = []
    for key, sections in photos.items():
        norm = normalize_period(key)
        if not norm:
            continue
        y, mo, ym = norm

        # 1순위: type ∈ {공정사진, 공사사진} 이면서 category=육상부
        # 2순위: type ∈ {공정사진, 공사사진} 인 모든 섹션
        # 3순위: 전 섹션
        priority_pools = [[], [], []]
        for sec in sections:
            t = sec.get("type")
            cat = sec.get("category")
            its = sec.get("items") or []
            if t in PROCESS_TYPES and cat in LAND_CATS:
                priority_pools[0].extend(its)
            if t in PROCESS_TYPES:
                priority_pools[1].extend(its)
            priority_pools[2].extend(its)

        pick = None
        for pool in priority_pools:
            pick = pick_panorama(pool)
            if pick:
                break
        if not pick:
            continue
        selected.append({
            "year": y, "month": mo, "ym": ym,
            "file": pick["file"],
            "caption": pick.get("caption", ""),
        })

    selected.sort(key=lambda r: (r["year"], r["month"]))
    return selected


# ---- 2) 공정현황 (2-1 시트) ------------------------------------------------

def parse_progress():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb["2-1.공정현황"]
    rows = []

    # 데이터는 5행부터 시작 (B열 = 공종, C=단위, D=수량, E=가중치, ...)
    # E = 가중치(EPC 전체 비율)
    # F = 실적 수량 (전월누계), G = 실적 수량 (금월), H = 실적 수량 (누계)
    # I = EPC 실적 진도 (전월누계), J = EPC 금월, K = EPC 누계(A)
    # L = EPC 계획(B), M = 계획대비 (A/B)
    TOP_LEVELS = {"ENGINEERING", "PROCUREMENT", "CONSTRUCTION"}

    cur_top = None
    cur_sub = None

    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        name_s = str(name).strip()
        if not name_s:
            continue

        unit = ws.cell(r, 3).value
        qty = ws.cell(r, 4).value
        weight = ws.cell(r, 5).value
        act_qty_cum = ws.cell(r, 8).value
        epc_actual = ws.cell(r, 11).value
        epc_plan = ws.cell(r, 12).value
        plan_vs = ws.cell(r, 13).value  # A/B

        # 분류: top / sub / item
        is_top = name_s in TOP_LEVELS
        is_sub = bool(re.match(r"^[IVX]+\.\s", name_s)) or (
            unit is None and not is_top and weight is not None
        )

        if is_top:
            cur_top = name_s
            cur_sub = None
            kind = "top"
        elif is_sub:
            cur_sub = name_s
            kind = "sub"
        else:
            kind = "item"

        # 자식 (들여쓰기 들어간 항목)은 cur_sub 안에 둔다
        is_child = bool(re.match(r"^\s+\d+\)\s", str(name)))

        rows.append({
            "name": name_s,
            "top": cur_top,
            "sub": cur_sub,
            "kind": "child" if is_child else kind,
            "unit": unit,
            "qty": float(qty) if isinstance(qty, (int, float)) else None,
            "weight": float(weight) if isinstance(weight, (int, float)) else None,
            "actual": float(epc_actual) if isinstance(epc_actual, (int, float)) else None,
            "plan": float(epc_plan) if isinstance(epc_plan, (int, float)) else None,
            "ratio": float(plan_vs) if isinstance(plan_vs, (int, float)) else None,
        })

    return {
        "title": "2-1. 공정현황",
        "source_sheet": "2-1.공정현황",
        "source_file": EXCEL.name,
        "rows": rows,
    }


# ---- 3) 주요Event (2-2 시트) ----------------------------------------------

def _norm_date(v, default_year=None):
    """엑셀 셀 → ISO 'YYYY-MM-DD' (가능하면). 실패시 None.
    셀이 이미 datetime이면 그대로, 'M/D' 같은 문자열이면 default_year 사용,
    'YYYY/MM/DD' 또는 'YYYY-MM-DD' 도 지원.
    """
    import datetime as dt
    if v is None: return None
    if isinstance(v, dt.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, dt.date):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s: return None
    # 'YYYY/MM/DD' 또는 'YYYY-MM-DD'
    m = re.match(r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$', s)
    if m:
        y, mo, da = map(int, m.groups())
        return f"{y:04d}-{mo:02d}-{da:02d}"
    # 'M/D' 또는 'MM/DD'
    m = re.match(r'^(\d{1,2})[/](\d{1,2})$', s)
    if m and default_year:
        mo, da = map(int, m.groups())
        return f"{default_year:04d}-{mo:02d}-{da:02d}"
    return None


def parse_events():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb['2-2.주요Event']
    events = []
    cur_year = None
    for r in range(5, ws.max_row + 1):
        no = ws.cell(r, 2).value
        # 연도 헤더 행 ('2022년' 등)
        if isinstance(no, str):
            m = re.match(r'^(\d{4})년', no)
            if m:
                cur_year = int(m.group(1))
            continue
        if no is None: continue
        cat = ws.cell(r, 4).value
        title = ws.cell(r, 5).value
        rev0 = ws.cell(r, 11).value
        rev6 = ws.cell(r, 17).value
        actual = ws.cell(r, 18).value
        note = ws.cell(r, 20).value
        if not title: continue
        events.append({
            'no': int(no),
            'year': cur_year,
            'category': (str(cat).strip() if cat else None),
            'title': str(title).strip(),
            'rev0': _norm_date(rev0, cur_year),
            'rev6': _norm_date(rev6, cur_year),
            'actual': _norm_date(actual, cur_year),
            'note': (str(note).strip() if note else None),
        })
    return {
        'title': '2-2. 주요 공정계획 및 실적',
        'source_sheet': '2-2.주요Event',
        'source_file': EXCEL.name,
        'events': events,
    }


# ---- 4) 품질관리이행실적 (6 시트) -----------------------------------------

def _float(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def parse_quality():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb['6.품질관리이행실적']

    # 6.1 텍스트 영역: 4월 실적 + 5월 계획 항목 (row 6~37 정도)
    # 단순 텍스트 리스트로 수집
    text_section = []
    for r in range(6, 38):
        cat = ws.cell(r, 2).value
        actual = ws.cell(r, 3).value  # 4월 실적
        plan_text = ws.cell(r, 12).value  # 5월 계획
        if actual or plan_text:
            text_section.append({
                'category': (str(cat).strip() if cat else None),
                'actual_apr': (str(actual).strip() if actual else None),
                'plan_may':   (str(plan_text).strip() if plan_text else None),
            })

    # 6.2 토목/건축공사 품질시험 실적 (행 42~60)
    civil_sections = []
    cur_section = None
    for r in range(42, 65):
        a = ws.cell(r, 2).value   # 구분 (대분류)
        b = ws.cell(r, 3).value   # 세부 항목
        unit = ws.cell(r, 4).value
        plan = ws.cell(r, 5).value
        # 총 누계 실적: 컬럼 13~16 (수행, 합격, 불합격, 불량률)
        cum_exec = _float(ws.cell(r, 13).value)
        cum_ok   = _float(ws.cell(r, 14).value)
        cum_ng   = _float(ws.cell(r, 15).value)
        cum_rej  = _float(ws.cell(r, 16).value)
        if a:
            cur_section = str(a).strip()
        if not b: continue
        civil_sections.append({
            'section': cur_section,
            'item': str(b).strip(),
            'unit': (str(unit).strip() if unit else None),
            'plan': plan if isinstance(plan, (int, float)) else None,
            'cum_exec': cum_exec, 'cum_ok': cum_ok, 'cum_ng': cum_ng, 'cum_reject_rate': cum_rej,
        })

    # 6.3 용접 불량률 (행 63~115): 여러 구분 영역
    # 일부 영역(운영지역 등)은 데이터가 A열(c=1)부터 시작하고,
    # 다른 영역은 B열(c=2)부터 시작 — 두 layout 모두 자동 감지
    weld_sections = []
    cur_region = None
    for r in range(63, 115):
        # 헤더 / 라벨 행 판별: A 또는 B 열에서 텍스트 확인
        a1 = ws.cell(r, 1).value
        a2 = ws.cell(r, 2).value
        item_text = None
        start_col = None  # 데이터의 시작 컬럼 (1 또는 2)
        for v, c in ((a1, 1), (a2, 2)):
            if v is not None:
                s = str(v).strip()
                if not s: continue
                if s.startswith('구분 ('):
                    cur_region = s.replace('구분 (', '').replace(')', '').strip()
                    item_text = '__HEADER__'
                    break
                if s in ('수행', '합격', '불합격', '불량율', '단위', '계획', '구분'):
                    item_text = '__SKIP__'
                    break
                item_text = s
                start_col = c
                break
        if not item_text or item_text in ('__HEADER__', '__SKIP__'):
            continue
        # 데이터 컬럼 위치: 시작 컬럼이 1이면 B/C부터 데이터, 2면 C/D부터.
        # 누계 4컬럼은 일관되게 '단위(c+2)', '계획(c+3)', 전월(c+4..c+7), 금월(c+8..c+11), 누계(c+12..c+15)
        # 하지만 우리는 그냥 누계 컬럼(c+12 ~ c+15) 만 필요
        base = start_col
        unit = ws.cell(r, base + 2).value
        cum_exec = _float(ws.cell(r, base + 12).value)
        cum_ok   = _float(ws.cell(r, base + 13).value)
        cum_ng   = _float(ws.cell(r, base + 14).value)
        cum_rej  = _float(ws.cell(r, base + 15).value)
        if cum_exec is None and cum_ok is None: continue
        weld_sections.append({
            'region': cur_region,
            'item': item_text,
            'unit': (str(unit).strip() if unit else None),
            'cum_exec': cum_exec, 'cum_ok': cum_ok, 'cum_ng': cum_ng, 'cum_reject_rate': cum_rej,
        })

    return {
        'title': '6. 품질관리이행실적',
        'source_sheet': '6.품질관리이행실적',
        'source_file': EXCEL.name,
        'text_items': text_section,
        'civil_tests': civil_sections,
        'welding': weld_sections,
    }


# ---- main -----------------------------------------------------------------

def main():
    pano = build_panorama()
    OUT_PANO.write_text(
        "window.PANORAMA_DATA = " + json.dumps(pano, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"[OK] panorama_data.js → {len(pano)} months")
    for r in pano[:3] + pano[-3:]:
        print(f"     - {r['ym']}: {r['caption']}")

    prog = parse_progress()
    OUT_PROG.write_text(
        "window.PROGRESS_STATUS_DATA = " + json.dumps(prog, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"[OK] progress_status_data.js → {len(prog['rows'])} rows")

    ev = parse_events()
    OUT_EVENT.write_text(
        "window.EVENTS_DATA = " + json.dumps(ev, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"[OK] events_data.js → {len(ev['events'])} events")

    qd = parse_quality()
    OUT_QUAL.write_text(
        "window.QUALITY_DATA = " + json.dumps(qd, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"[OK] quality_data.js → text {len(qd['text_items'])} · civil {len(qd['civil_tests'])} · welding {len(qd['welding'])}")


if __name__ == "__main__":
    main()
