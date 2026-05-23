"""
extract_construction_photos.py

25.1월 ~ 26.5월 월간안전공정회의 엑셀 파일들의
11.공사사진대장 시트에서 이미지+캡션을 추출하여,
기존 photos.js (PHOTOS_DATA) 에 누락된 항목만 추가한다.

매칭 규칙
- 캡션 셀: (anchor.row + 1 + 12, anchor.col + 1)  (1-indexed)
  즉, 이미지 anchor 의 행 + 12 (행 13칸 아래), 그리고 anchor col + 1 (왼쪽 위 모서리)
- 카테고리 (공사사진대장 구분 헤더):
    토목: row 3~54, 건축: 55~80, 기계: 81~132, 전기: 133~
- 모든 카테고리를 category='육상부', type='공사사진' 섹션으로 매핑
  (탱크/현장 공사 사진이므로 — 154kV 별도 substation 사진은 11번 시트에 거의 없음)
  섹션 title 에 디시플린 명시: "공사 사진 (토목) - 2026.05월"

중복 방지
- 캡션 + 월 키로 비교. 정확히 같은 캡션이 같은 월에 있으면 skip.
- 사진 바이너리는 SHA-1 해시로 중복 검사하여 동일 파일은 이미 있던 경로 재사용.
"""
import hashlib
import io
import json
import re
import sys
import zipfile
from pathlib import Path

import openpyxl

# 콘솔 UTF-8 강제 (Windows cp949 에서도 em-dash 등 출력 가능하게)
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).parent
PHOTOS_DIR = ROOT / "photos"
PHOTOS_JS = ROOT / "photos.js"

# 처리할 엑셀 파일 (월별)
EXCEL_FILES = [
    # 24.1월 ~ 24.11월: Fasoo DRM 보호로 openpyxl 불가 → 스킵
    ("2024-12", "광)7,8탱크_월간안전공정회의_24.12월_Final.xlsx"),
    ("2025-01", "광)LNG #7,8탱크_월간안전공정회의_25.1월_Final_R3.xlsx"),
    ("2025-02", "광)LNG #7,8탱크_월간안전공정회의_25.2월_Final_R1.xlsx"),
    ("2025-03", "광)LNG #7,8탱크_월간안전공정회의_25.03월_Final_R1.xlsx"),
    ("2025-04", "광)LNG #7,8탱크_월간안전공정회의_25.04월_Final.xlsx"),
    ("2025-05", "터미널건설추진반_광) LNG 7,8탱크_월간안전공정회의_25.05월_Final_R2_20250513_김중혁(j-hyeok.kim)_사외비A.xlsx"),
    ("2025-06", "광)LNG #7,8탱크_월간안전공정회의_25.6월_Final_R1.xlsx"),
    ("2025-07", "광)LNG #7,8탱크_월간안전공정회의_25.7월_Final.xlsx"),
    ("2025-08", "광)LNG #7,8탱크_월간안전공정회의_25.8월_Final_R1.xlsx"),
    ("2025-09", "광)LNG #7,8탱크_월간안전공정회의_25.9월_Final.xlsx"),
    ("2025-10", "광)LNG 7,8탱크_월간안전공정회의_25.10월_Final.xlsx"),
    ("2025-11", "광)LNG #7,8탱크_월간안전공정회의_25.11월_Final.xlsx"),
    ("2025-12", "광)LNG #7,8탱크_월간안전공정회의_25.12월_Final_R1.xlsx"),
    ("2026-01", "광)LNG #7,8탱크_월간안전공정회의_26.1월_Final.xlsx"),
    ("2026-02", "광)LNG #7,8탱크_월간안전공정회의_26.2월_Final_R1.xlsx"),
    ("2026-03", "광)LNG #7,8탱크_월간안전공정회의_26.3월_Final.xlsx"),
    ("2026-04", "광)LNG #7,8탱크_월간안전공정회의_26.4월_Final.xlsx"),
    ("2026-05", "광)LNG #7,8탱크_월간안전공정회의_26.5월_Final.xlsx"),
]

# 카테고리 헤더 → row 범위 매핑 (실시간 추출됨)
def find_discipline_for_row(row_1idx, headers):
    """row_1idx 가 어느 카테고리에 속하는지 결정. headers = [(row, disc), ...]"""
    cur = None
    for hrow, disc in headers:
        if hrow <= row_1idx:
            cur = disc
        else:
            break
    return cur or "기타"


def detect_headers(ws):
    """B열에서 카테고리 헤더 (토목/건축/기계/전기 등) 추출"""
    headers = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if not v or not isinstance(v, str):
            continue
        s = v.strip()
        if len(s) > 10:
            continue
        if "/" in s or "월" in s or ":" in s or "(" in s:
            continue
        # 단순 공백 제거 후 정규화
        norm = s.replace(" ", "")
        if norm in ("토목", "건축", "기계", "전기"):
            headers.append((r, norm))
    return headers


def normalize_disc(name):
    return {
        "토목": "토목",
        "건축": "건축",
        "기계": "기계",
        "전기": "전기",
    }.get(name, "기타")


def load_photos_data():
    text = PHOTOS_JS.read_text(encoding="utf-8")
    text = re.sub(r"^\s*window\.PHOTOS_DATA\s*=\s*", "", text, count=1)
    text = text.rstrip().rstrip(";")
    return json.loads(text)


def save_photos_data(data):
    js = "window.PHOTOS_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n"
    PHOTOS_JS.write_text(js, encoding="utf-8")


def get_image_bytes(xlsx_path, image_ref):
    """xlsx 의 zip 에서 image_ref (예: '/xl/media/image1.jpeg' 또는 'xl/media/image1.jpeg') 의 바이너리를 반환"""
    ref = image_ref.lstrip("/")
    with zipfile.ZipFile(xlsx_path, "r") as z:
        return z.read(ref)


def main():
    photos = load_photos_data()
    added_total = 0
    skipped_dup = 0
    skipped_no_caption = 0

    # 사진 binary hash → 기존 파일 경로 캐시 (중복 binary 재사용)
    binary_cache = {}

    converted_dirs = [ROOT / "_xlsx_converted_all", ROOT / "_xlsx_converted"]
    for ym, fname in EXCEL_FILES:
        # 원본이 정상 zip(.xlsx)이면 원본 우선, 아니면 변환본 (헤더가 정상인 것) 사용
        primary = ROOT / fname
        path = None

        def _is_zip(p):
            try:
                return open(p, "rb").read(4).startswith(b"PK\x03\x04")
            except Exception:
                return False

        if primary.exists() and _is_zip(primary):
            path = primary
        else:
            for d in converted_dirs:
                cand = d / fname
                if cand.exists() and _is_zip(cand):
                    path = cand
                    break
            if path is None and primary.exists():
                path = primary  # 마지막 fallback (어차피 BadZipFile 로 skip 됨)
        if path is None:
            print(f"[SKIP] {ym}: 파일 없음 ({fname})")
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except zipfile.BadZipFile:
            print(f"[SKIP] {ym}: OLE2(.xls) 헤더 - openpyxl 불가 ({path})")
            continue
        except Exception as e:
            print(f"[ERR ] {ym}: 엑셀 로드 실패 - {e}")
            continue
        if "11.공사사진대장" not in wb.sheetnames:
            print(f"[SKIP] {ym}: 11.공사사진대장 시트 없음")
            continue
        ws = wb["11.공사사진대장"]
        imgs = list(getattr(ws, "_images", []))
        if not imgs:
            print(f"[INFO] {ym}: 임베디드 이미지 없음")
            continue

        headers = detect_headers(ws)

        # 이미 photos.js 에 있는 캡션 목록 (이 월에 한해)
        existing_captions = set()
        if ym in photos:
            for sec in photos[ym]:
                for it in sec.get("items") or []:
                    cap = (it.get("caption") or "").strip()
                    if cap:
                        existing_captions.add(cap)

        # 디스크에 저장할 폴더 준비
        target_dir = PHOTOS_DIR / ym
        target_dir.mkdir(parents=True, exist_ok=True)

        added_for_month = 0
        per_disc_items = {"토목": [], "건축": [], "기계": [], "전기": [], "기타": []}

        for im in imgs:
            anc = im.anchor
            try:
                ar = anc._from.row  # 0-indexed
                ac = anc._from.col  # 0-indexed
            except Exception:
                continue
            # 캡션 위치 = anchor.row + 1 (1-indexed 시작) + 12 = anchor.row + 13
            caption_row = ar + 13
            caption_col = ac + 1  # 1-indexed
            caption = None
            try:
                v = ws.cell(caption_row, caption_col).value
                if v and isinstance(v, str):
                    caption = v.strip()
            except Exception:
                caption = None

            # 다른 위치 후보도 검사 (캡션이 1행 위/아래 있을 수도 있음)
            if not caption:
                for dr in (-1, 1, 12, 14):
                    try:
                        v = ws.cell(ar + 1 + dr, ac + 1).value
                        if v and isinstance(v, str) and len(v.strip()) > 5:
                            caption = v.strip()
                            break
                    except Exception:
                        continue

            if not caption:
                skipped_no_caption += 1
                continue

            # 중복 체크
            if caption in existing_captions:
                skipped_dup += 1
                continue
            existing_captions.add(caption)

            # 이미지 바이너리 추출 — openpyxl Image._data() 사용 (BytesIO 안전)
            data = None
            try:
                if callable(getattr(im, "_data", None)):
                    data = im._data()
                elif hasattr(im, "ref"):
                    ref_obj = im.ref
                    if hasattr(ref_obj, "read"):
                        ref_obj.seek(0)
                        data = ref_obj.read()
                    elif isinstance(ref_obj, str):
                        data = get_image_bytes(path, ref_obj)
            except Exception as e:
                print(f"  [ERR] {ym} 이미지 추출 실패: {e}")
                continue
            if not data:
                continue

            # 해시 기반 중복 binary 캐시 (같은 binary는 동일 파일 재사용)
            h = hashlib.sha1(data).hexdigest()[:12]
            # 확장자: 시그니처로 판별
            if data[:3] == b"\xff\xd8\xff":
                ext = ".jpg"
            elif data[:8] == b"\x89PNG\r\n\x1a\n":
                ext = ".png"
            elif data[:6] in (b"GIF87a", b"GIF89a"):
                ext = ".gif"
            else:
                ext = ".jpg"

            disc = normalize_disc(find_discipline_for_row(ar + 1, headers))
            # 파일명: cnstr_{disc}_{hash}{ext}
            fname_out = f"cnstr_{disc}_{h}{ext}"
            out_path = target_dir / fname_out
            if not out_path.exists():
                out_path.write_bytes(data)

            rel = f"photos/{ym}/{fname_out}"
            per_disc_items[disc].append({"file": rel, "caption": caption})
            added_for_month += 1

        # photos.js 에 병합
        if added_for_month > 0:
            if ym not in photos:
                photos[ym] = []
            for disc, items in per_disc_items.items():
                if not items:
                    continue
                # 동일 (title, category, type) 섹션 있으면 그쪽에 append, 없으면 신규
                # title 패턴: "공사 사진 ({disc}) - {ym 라벨}월"
                ym_lbl = ym.replace("-", ".")
                section_title = f"공사 사진 ({disc}) — {ym_lbl}"
                target_sec = None
                for sec in photos[ym]:
                    if (sec.get("title") == section_title
                            and sec.get("type") == "공사사진"
                            and sec.get("category") == "육상부"):
                        target_sec = sec
                        break
                if not target_sec:
                    target_sec = {
                        "title": section_title,
                        "category": "육상부",
                        "type": "공사사진",
                        "items": [],
                        "slide": 11,
                    }
                    photos[ym].append(target_sec)
                target_sec["items"].extend(items)

        print(f"[OK ] {ym}: +{added_for_month} 추가 / {len(imgs)} 이미지 (헤더 {len(headers)}개)")
        added_total += added_for_month

    save_photos_data(photos)
    print()
    print(f"=== 완료: 총 추가 {added_total} · 중복 캡션 스킵 {skipped_dup} · 캡션 없음 스킵 {skipped_no_caption} ===")


if __name__ == "__main__":
    main()
